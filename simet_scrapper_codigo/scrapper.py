# -*- coding: utf-8 -*-
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import csv
import time
from openpyxl import load_workbook
from openpyxl import Workbook
from numpy import median


base_table = load_workbook('worktable.xlsx')
base_table_sheet = base_table['Planilha1']

firefox = webdriver.Firefox()
# Página web com os dados do SIMET
firefox.get('http://simet-publico.ceptro.br/mapas7/')

# Espera um pouco até que a página seja carregada, caso contrário é possível que alguns dos elementos não sejam encontrados
time.sleep(5)

iterator = 585

for iterator_base_table in range(585, 593):
    iterator += 1

    base_table_codigoINEP = str(base_table_sheet.cell(row = iterator_base_table, column = 4).value)
    # Busca pelo inputText de busca, aqui é onde vão ser inseridos os endereços lat, lon
    enderecoInputTextField = firefox.find_element_by_xpath('//*[@id="searchBox"]/input')
    enderecoInputTextField.send_keys(base_table_codigoINEP)
    # Simula o botão enter
    enderecoInputTextField.send_keys(Keys.ENTER)

    # Espera um pouco até que a página seja carregada, caso contrário é possível que alguns dos elementos não sejam encontrados
    if iterator == 1:
        time.sleep(10)
    else:
        time.sleep(5)

    mensagemDeResultado = firefox.find_element_by_class_name('highcharts-title').text

    print(mensagemDeResultado == "Todos ASN")

    if (mensagemDeResultado == "Todos ASN"):
        #dialogWindow = firefox.find_element_by_xpath('//*[@id="highcharts-0"]/svg/text[2]')
        qtdDeMedicoesNaRegiao = firefox.find_element_by_class_name('highcharts-subtitle').text.split()[0]
        base_table_sheet.cell(row=iterator_base_table, column = 8).value = qtdDeMedicoesNaRegiao
        #print("debug 1")

        # Espera um pouco até que a página seja carregada, caso contrário é possível que alguns dos elementos não sejam encontrados
        time.sleep(1)

        listaDeASN = firefox.find_element_by_id("asnButtonContainer").text.splitlines()[1:]

        qtdOperadoraSIMET = len(listaDeASN)
        base_table_sheet.cell(row=iterator_base_table, column = 9).value = qtdOperadoraSIMET
        #print("debug 2")

        velocidadesMediasDasOperadoras = []
        somaVelocidadesMediasDasOperadoras = 0
        nomeDasOperadoras = ""

        for item in listaDeASN:
            velocidadesMediasDasOperadoras.append(float(item.split()[-2]))
            somaVelocidadesMediasDasOperadoras += float(item.split()[-2])
            nomeDasOperadoras += item.split('-')[1][1:] + ", "
            # asnDaOperadora = item.split("-")[0]
            # nomeDaOperadora = item.split('-')[1][1:]
            # velocidadeMedia = item.split()[-2]

        base_table_sheet.cell(row=iterator_base_table, column = 10).value = nomeDasOperadoras
        base_table_sheet.cell(row=iterator_base_table, column = 11).value = somaVelocidadesMediasDasOperadoras/len(velocidadesMediasDasOperadoras)
        base_table_sheet.cell(row=iterator_base_table, column = 12).value = max(velocidadesMediasDasOperadoras)
        base_table_sheet.cell(row=iterator_base_table, column = 13).value = min(velocidadesMediasDasOperadoras)

        tabelaComVelocidades = firefox.find_elements_by_xpath('//*[name()="svg"]/*[name()="g"]')

        grafico = tabelaComVelocidades[4].find_elements_by_tag_name("rect")

        index = 0

        for item in grafico:
            if float(item.get_attribute("height")) > 0:
                base_table_sheet.cell(row=iterator_base_table, column = 14+index).value = 1
            else:
                base_table_sheet.cell(row=iterator_base_table, column = 14+index).value = 0
            index += 1
            time.sleep(1)

        webdriver.ActionChains(firefox).send_keys(Keys.ESCAPE).perform()

    else:
        #dialogWindow = firefox.find_element_by_xpath('//*[@id="highcharts-0"]/svg/text[2]')
        base_table_sheet.cell(row=iterator_base_table, column = 8).value = 0
        #print("debug 1")

        base_table_sheet.cell(row=iterator_base_table, column = 9).value = 0
        #print("debug 2")

        base_table_sheet.cell(row=iterator_base_table, column = 10).value = 0
        base_table_sheet.cell(row=iterator_base_table, column = 11).value = 0
        base_table_sheet.cell(row=iterator_base_table, column = 12).value = 0
        base_table_sheet.cell(row=iterator_base_table, column = 13).value = 0

        for iterator2 in range(14, 29):
            base_table_sheet.cell(row=iterator_base_table, column = iterator2).value = 0

        webdriver.ActionChains(firefox).send_keys(Keys.ESCAPE).perform()

    print("Iterator numero: " + str(iterator))
    base_table.save("worktable.xlsx")
